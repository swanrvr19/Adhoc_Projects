"""
Build the claims restatement workbook pulling ultimate_incurred from each of the
three completed dashboards' 'combined data' tabs (202511, 202512, 202601) and
showing restatement across report dates.
"""
import os
import pandas as pd
import datetime as dt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/sessions/magical-bold-goodall/mnt/Medicare_Dashboard/Claude Cowork for Actuaries - Lesson 01/Guided"
OUT = f"{BASE}/Claims Restatement - All Runs.xlsx"

REPORT_DATES = [
    ('202511', dt.datetime(2025, 11, 30)),
    ('202512', dt.datetime(2025, 12, 31)),
    ('202601', dt.datetime(2026, 1, 31)),
]

# ---- Read combined data (claims rows) from each workbook
def read_combined_claims(report_date):
    p = f"{BASE}/{report_date}/Claude Cowork for Actuaries - Claims Dashboard - {report_date}.xlsx"
    wb = load_workbook(p, data_only=True)
    ws = wb['combined data']
    records = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        year, month, plan, claim, mbr_, ult = row[0], row[1], row[2], row[3], row[4], row[5]
        if year is None: continue
        if claim is None: continue  # skip mbr rows
        records.append({
            'report_date': report_date,
            'incurred_year': int(year),
            'incurred_month': int(month),
            'plan_type': plan,
            'claim_type': claim,
            'ultimate_incurred': float(ult) if ult is not None else 0.0,
        })
    return pd.DataFrame(records)

frames = []
for rd, _ in REPORT_DATES:
    df = read_combined_claims(rd)
    print(f"{rd}: {len(df)} claim rows, through {df.incurred_year.max()}-{df[df.incurred_year==df.incurred_year.max()].incurred_month.max()}")
    frames.append(df)
combined = pd.concat(frames, ignore_index=True)

# Add incurred_date (first of month style, stored as period string YYYYMM)
combined['incurred_ym'] = combined['incurred_year'] * 100 + combined['incurred_month']

# Compute lag at each report date
def lag_from(report_yyyymm, inc_yyyymm):
    ry, rm = int(str(report_yyyymm)[:4]), int(str(report_yyyymm)[4:])
    iy, im = inc_yyyymm // 100, inc_yyyymm % 100
    return (ry - iy) * 12 + (rm - im)

combined['lag'] = combined.apply(
    lambda r: lag_from(r['report_date'], r['incurred_ym']), axis=1
)

# Only keep rows where lag >= 2 (dashboard convention)
restate = combined[combined['lag'] >= 2].copy()
print(f"\nTotal rows at lag 2+: {len(restate)}")
print(f"Incurred months: {sorted(restate['incurred_ym'].unique())}")

# ---- Build Tab 1: Restatement Grid (total by incurred month)
grid_total = restate.groupby(['incurred_ym', 'report_date'])['ultimate_incurred'].sum().unstack('report_date')
grid_total = grid_total.reindex(sorted(restate['incurred_ym'].unique()))
grid_total = grid_total[['202511', '202512', '202601']]

# ---- Build Tab 2: Restatement by Segment
restate['segment'] = restate['plan_type'] + ' / ' + restate['claim_type']
grid_seg = restate.groupby(['segment', 'incurred_ym', 'report_date'])['ultimate_incurred'].sum().unstack('report_date')
grid_seg = grid_seg.reindex(columns=['202511', '202512', '202601'])

# ---- Utility: month label (YYYYMM -> 'Jan 2024')
def ym_label(yyyymm):
    y = yyyymm // 100
    m = yyyymm % 100
    return dt.date(y, m, 1).strftime('%b %Y')

# ========================================================================
# BUILD WORKBOOK
# ========================================================================
wb = Workbook()

# Styles
navy = 'FF1F3A5F'
light = 'FFE8EDF3'
yellow_hi = 'FFFFF2CC'
red_hi = 'FFF8CBAD'
green_hi = 'FFC6EFCE'

title_font = Font(name='Calibri', bold=True, size=14, color='FFFFFFFF')
header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFFFF')
subheader_font = Font(name='Calibri', bold=True, size=10, color=navy)
body_font = Font(name='Calibri', size=10)
italic_font = Font(name='Calibri', italic=True, size=9, color='FF555555')

fill_navy = PatternFill('solid', fgColor=navy)
fill_light = PatternFill('solid', fgColor=light)
fill_yellow = PatternFill('solid', fgColor=yellow_hi)
fill_red = PatternFill('solid', fgColor=red_hi)
fill_green = PatternFill('solid', fgColor=green_hi)

thin_side = Side(style='thin', color='FFB0B7C3')
cell_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

dollar_fmt = '_("$"* #,##0_);_("$"* (#,##0);_("$"* "-"??_);_(@_)'
pct_fmt = '0.0%;(0.0%);-'
change_dollar_fmt = '_("$"* #,##0_);[Red]_("$"* (#,##0);_("$"* "-"??_);_(@_)'

def write_grid(ws, row_labels, data_values, start_row, start_col, label_header='Incurred Month'):
    """
    Write a restatement grid with columns:
    Incurred | Lag@1stReport | 202511 | 202512 | 202601 | Δ Nov→Dec | % | Δ Dec→Jan | %
    data_values: dict of {row_label: {'202511': val, '202512': val, '202601': val, 'lag_start': X}}
    Returns (last_row, last_col)
    """
    headers = [label_header, 'Lag at earliest', '202511 Ultimate', '202512 Ultimate', '202601 Ultimate',
               '$ Change 202511→202512', '% Change 202511→202512',
               '$ Change 202512→202601', '% Change 202512→202601']
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=start_col + j, value=h)
        c.font = header_font
        c.fill = fill_navy
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = cell_border
    ws.row_dimensions[start_row].height = 40

    r = start_row + 1
    for label in row_labels:
        vals = data_values[label]
        v511 = vals.get('202511')
        v512 = vals.get('202512')
        v601 = vals.get('202601')
        lag_start = vals.get('lag_start')

        row_cells = [
            ws.cell(row=r, column=start_col + 0, value=label),
            ws.cell(row=r, column=start_col + 1, value=lag_start),
            ws.cell(row=r, column=start_col + 2, value=v511),
            ws.cell(row=r, column=start_col + 3, value=v512),
            ws.cell(row=r, column=start_col + 4, value=v601),
        ]

        # Changes only when both adjacent values exist
        d1 = v512 - v511 if (v511 is not None and v512 is not None) else None
        p1 = (d1 / v511) if (d1 is not None and v511 not in (None, 0)) else None
        d2 = v601 - v512 if (v512 is not None and v601 is not None) else None
        p2 = (d2 / v512) if (d2 is not None and v512 not in (None, 0)) else None

        row_cells.append(ws.cell(row=r, column=start_col + 5, value=d1))
        row_cells.append(ws.cell(row=r, column=start_col + 6, value=p1))
        row_cells.append(ws.cell(row=r, column=start_col + 7, value=d2))
        row_cells.append(ws.cell(row=r, column=start_col + 8, value=p2))

        # Format
        for j, cell in enumerate(row_cells):
            cell.font = body_font
            cell.border = cell_border
            if j == 0:
                cell.alignment = Alignment(horizontal='left')
            elif j == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='right')
            if j in (2, 3, 4, 5, 7):
                cell.number_format = dollar_fmt
            elif j in (6, 8):
                cell.number_format = pct_fmt

        # Highlight >5% movements
        for pct_cell, pct_val in [(row_cells[6], p1), (row_cells[8], p2)]:
            if pct_val is not None and abs(pct_val) > 0.05:
                if pct_val > 0:
                    pct_cell.fill = fill_yellow  # adverse development (up)
                else:
                    pct_cell.fill = fill_green   # favorable development (down)
        r += 1
    return r - 1, start_col + len(headers) - 1

def earliest_lag(yyyymm, report_date_list):
    """First report_date where incurred month is at lag >= 2."""
    for rd in report_date_list:
        l = lag_from(rd, yyyymm)
        if l >= 2:
            return l, rd
    return None, None

# ========================================================================
# Tab 1: Restatement Grid
# ========================================================================
ws1 = wb.active
ws1.title = 'Restatement Grid'

ws1.column_dimensions['A'].width = 2
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 17
ws1.column_dimensions['E'].width = 17
ws1.column_dimensions['F'].width = 17
ws1.column_dimensions['G'].width = 15
ws1.column_dimensions['H'].width = 15
ws1.column_dimensions['I'].width = 15
ws1.column_dimensions['J'].width = 15

ws1.merge_cells('B2:J2')
ws1['B2'] = 'Claims Restatement — Ultimate Incurred by Report Date'
ws1['B2'].font = Font(name='Calibri', bold=True, size=14, color=navy)
ws1['B2'].alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B3:J3')
ws1['B3'] = ('Rows: incurred months. Columns: report dates (Nov 2025, Dec 2025, Jan 2026 closes). '
             'Shows movement between consecutive report dates. Completion dashboards exclude '
             'lag 0 and 1 — first observable restatement is lag 2 → lag 3.')
ws1['B3'].font = italic_font
ws1['B3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws1.row_dimensions[3].height = 32

# Build rows
data = {}
row_labels = []
report_dates = [rd for rd, _ in REPORT_DATES]

# Use all incurred months that appear at lag 2+ in ANY report
all_ym = sorted(restate['incurred_ym'].unique())

for ym in all_ym:
    label = ym_label(ym)
    row_labels.append(label)
    vals = {}
    for rd in report_dates:
        if (ym, rd) in grid_total.stack().index if isinstance(grid_total, pd.DataFrame) else False:
            vals[rd] = grid_total.loc[ym, rd]
    # simpler
    row_data = grid_total.loc[ym]
    for rd in report_dates:
        v = row_data.get(rd)
        if pd.notna(v):
            # But only if lag >= 2 at that report date
            if lag_from(rd, ym) >= 2:
                vals[rd] = float(v)
            else:
                vals[rd] = None
        else:
            vals[rd] = None
    lag_start, _ = earliest_lag(ym, report_dates)
    vals['lag_start'] = lag_start
    data[label] = vals

last_row, last_col = write_grid(ws1, row_labels, data, start_row=5, start_col=2)

# Footer / legend
legend_row = last_row + 2
ws1.cell(row=legend_row, column=2, value='Highlights:').font = subheader_font
ws1.cell(row=legend_row + 1, column=2, value='  Yellow = adverse development (>+5%)').fill = fill_yellow
ws1.cell(row=legend_row + 2, column=2, value='  Green  = favorable development (<−5%)').fill = fill_green
ws1.cell(row=legend_row + 3, column=2, value='  Blank = incurred month not yet at lag 2 for that report')
for rr in range(legend_row, legend_row + 4):
    ws1.cell(row=rr, column=2).font = body_font

ws1.freeze_panes = 'C6'

# ========================================================================
# Tab 2: Restatement by Segment
# ========================================================================
ws2 = wb.create_sheet('Restatement by Segment')

ws2.column_dimensions['A'].width = 2
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 14
for col_letter in ['D', 'E', 'F']:
    ws2.column_dimensions[col_letter].width = 17
for col_letter in ['G', 'H', 'I', 'J']:
    ws2.column_dimensions[col_letter].width = 15

ws2.merge_cells('B2:J2')
ws2['B2'] = 'Claims Restatement — By Plan Type & Claim Type'
ws2['B2'].font = Font(name='Calibri', bold=True, size=14, color=navy)
ws2['B2'].alignment = Alignment(horizontal='left', vertical='center')

ws2.merge_cells('B3:J3')
ws2['B3'] = ('Four segments: HMO/IP, HMO/Non-IP, PPO/IP, PPO/Non-IP. Shows where development '
             'is concentrated — by book of business and by service category.')
ws2['B3'].font = italic_font
ws2['B3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[3].height = 28

segments = ['HMO / IP', 'HMO / Non-IP', 'PPO / IP', 'PPO / Non-IP']
cur_row = 5
for seg in segments:
    # Section header
    ws2.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=10)
    sc = ws2.cell(row=cur_row, column=2, value=seg)
    sc.font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
    sc.fill = PatternFill('solid', fgColor=navy)
    sc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws2.row_dimensions[cur_row].height = 22
    cur_row += 1

    seg_data = {}
    seg_labels = []
    seg_df = grid_seg.loc[seg] if seg in grid_seg.index.get_level_values(0) else None

    for ym in all_ym:
        label = ym_label(ym)
        seg_labels.append(label)
        vals = {}
        if seg_df is not None and ym in seg_df.index:
            row_data = seg_df.loc[ym]
            for rd in report_dates:
                v = row_data.get(rd)
                if pd.notna(v) and lag_from(rd, ym) >= 2:
                    vals[rd] = float(v)
                else:
                    vals[rd] = None
        else:
            for rd in report_dates:
                vals[rd] = None
        lag_start, _ = earliest_lag(ym, report_dates)
        vals['lag_start'] = lag_start
        seg_data[label] = vals

    cur_row, _ = write_grid(ws2, seg_labels, seg_data, start_row=cur_row, start_col=2)
    cur_row += 2  # spacer between sections

ws2.freeze_panes = 'C6'

# ========================================================================
# Tab 3: Summary of Findings
# ========================================================================
ws3 = wb.create_sheet('Summary of Findings')
ws3.column_dimensions['A'].width = 2
ws3.column_dimensions['B'].width = 110

ws3.merge_cells('B2:B2')
ws3['B2'] = 'Claims Restatement — Summary of Findings'
ws3['B2'].font = Font(name='Calibri', bold=True, size=14, color=navy)

# Compute observations dynamically
def biggest_movement(grid_df, report_from, report_to):
    deltas = []
    for ym in grid_df.index:
        a = grid_df.loc[ym, report_from]
        b = grid_df.loc[ym, report_to]
        if pd.notna(a) and pd.notna(b) and lag_from(report_from, ym) >= 2 and lag_from(report_to, ym) >= 2:
            if a != 0:
                pct = (b - a) / a
                deltas.append((ym, a, b, b - a, pct))
    return sorted(deltas, key=lambda x: abs(x[4]), reverse=True)

mv_nov_dec = biggest_movement(grid_total, '202511', '202512')
mv_dec_jan = biggest_movement(grid_total, '202512', '202601')

# Segment-level analysis (aggregate dollar change by segment)
def segment_change_totals(grid_seg, report_from, report_to):
    out = {}
    for seg in grid_seg.index.get_level_values(0).unique():
        seg_df = grid_seg.loc[seg]
        total_delta = 0
        for ym in seg_df.index:
            a = seg_df.loc[ym].get(report_from)
            b = seg_df.loc[ym].get(report_to)
            if pd.notna(a) and pd.notna(b) and lag_from(report_from, ym) >= 2 and lag_from(report_to, ym) >= 2:
                total_delta += (b - a)
        out[seg] = total_delta
    return out

seg_nov_dec = segment_change_totals(grid_seg, '202511', '202512')
seg_dec_jan = segment_change_totals(grid_seg, '202512', '202601')

def fmt_usd(x):
    return f"${x:,.0f}" if x >= 0 else f"(${abs(x):,.0f})"

bullets = []
bullets.append('Observations')
bullets.append('')

# Biggest movements Nov→Dec
bullets.append('Largest movements from the Nov 2025 close to the Dec 2025 close:')
for ym, a, b, d, p in mv_nov_dec[:5]:
    lag_a = lag_from('202511', ym)
    lag_b = lag_from('202512', ym)
    bullets.append(f"   • {ym_label(ym)} (lag {lag_a} → {lag_b}): {fmt_usd(d)}  ({p:+.1%})")
bullets.append('')

# Biggest movements Dec→Jan
bullets.append('Largest movements from the Dec 2025 close to the Jan 2026 close:')
for ym, a, b, d, p in mv_dec_jan[:5]:
    lag_a = lag_from('202512', ym)
    lag_b = lag_from('202601', ym)
    bullets.append(f"   • {ym_label(ym)} (lag {lag_a} → {lag_b}): {fmt_usd(d)}  ({p:+.1%})")
bullets.append('')

# Segment-level development
bullets.append('Where development is concentrated (dollar change across all months with observable restatement):')
bullets.append('')
bullets.append('  Nov → Dec:')
for seg in sorted(seg_nov_dec, key=lambda s: abs(seg_nov_dec[s]), reverse=True):
    bullets.append(f"   • {seg}: {fmt_usd(seg_nov_dec[seg])}")
bullets.append('')
bullets.append('  Dec → Jan:')
for seg in sorted(seg_dec_jan, key=lambda s: abs(seg_dec_jan[s]), reverse=True):
    bullets.append(f"   • {seg}: {fmt_usd(seg_dec_jan[seg])}")
bullets.append('')

# Anomalies
bullets.append('Anomalies / things to watch:')
bullets.append('')
bullets.append('  • Lag 2 → lag 3 is the first observable restatement window. Movements here are expected because')
bullets.append('    at lag 2 the completed estimate still leans heavily on the factor; the lag 3 value reflects')
bullets.append('    an additional month of paid claims development.')
bullets.append('  • Watch for directionally consistent movements across months in a given segment — that would')
bullets.append('    suggest the underlying completion factors for that segment are systematically off.')
bullets.append('  • Mature months (lag 10+) should be essentially flat across reports, since CFs are 1.000 at')
bullets.append('    those lags. Any movement there would indicate a data correction upstream, not true development.')
bullets.append('  • Lags 0 and 1 are excluded by design. The dashboard does not display them because paid claims')
bullets.append('    at those lags are too thin for the completed estimate to be reliable. This grid respects that')
bullets.append('    convention — restatement starts at lag 2 → 3.')
bullets.append('')

# Headline readable observations (will depend on data)
bullets.append('Key takeaways:')
bullets.append('')
# Find whether IP or Non-IP is bigger mover
total_ip_movement = sum(v for k, v in seg_dec_jan.items() if 'IP' in k and 'Non-IP' not in k)
total_nonip_movement = sum(v for k, v in seg_dec_jan.items() if 'Non-IP' in k)
total_hmo_movement = sum(v for k, v in seg_dec_jan.items() if k.startswith('HMO'))
total_ppo_movement = sum(v for k, v in seg_dec_jan.items() if k.startswith('PPO'))

if abs(total_ip_movement) > abs(total_nonip_movement):
    bullets.append(f"  • IP development ({fmt_usd(total_ip_movement)}) is the bigger mover than Non-IP ({fmt_usd(total_nonip_movement)}) in the latest run.")
else:
    bullets.append(f"  • Non-IP development ({fmt_usd(total_nonip_movement)}) is the bigger mover than IP ({fmt_usd(total_ip_movement)}) in the latest run.")

if abs(total_hmo_movement) > abs(total_ppo_movement):
    bullets.append(f"  • HMO development ({fmt_usd(total_hmo_movement)}) exceeds PPO ({fmt_usd(total_ppo_movement)}) in the latest run.")
else:
    bullets.append(f"  • PPO development ({fmt_usd(total_ppo_movement)}) exceeds HMO ({fmt_usd(total_hmo_movement)}) in the latest run.")

# Write bullets to sheet
r = 4
for line in bullets:
    c = ws3.cell(row=r, column=2, value=line)
    if line in ('Observations', 'Key takeaways:', 'Anomalies / things to watch:',
                'Where development is concentrated (dollar change across all months with observable restatement):') \
        or line.startswith('Largest movements'):
        c.font = subheader_font
    else:
        c.font = body_font
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    r += 1

ws3.row_dimensions[2].height = 22

# ---- Save
wb.save(OUT)
print(f"\nWrote: {OUT}")
