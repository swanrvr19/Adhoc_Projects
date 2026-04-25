"""
Build the January 2026 dashboard by updating the copied Dec template in place.
- Update controls!B5 close date
- Replace claims data (rows 7+) with Jan26 CSV; extend formulas E:I
- Replace mbr data (rows 7+) with Jan26 CSV; extend formulas D:E
- Recompute pivot cached values (year x month aggregates) for 2024 and 2025
- Rebuild combined data with 25-month history
- Save; recalc via LibreOffice
"""
import os, sys, shutil, datetime as dt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

BASE = "/sessions/magical-bold-goodall/mnt/Medicare_Dashboard/Claude Cowork for Actuaries - Lesson 01/Guided"
JAN_FILE = f"{BASE}/202601/Claude Cowork for Actuaries - Claims Dashboard - 202601.xlsx"
CLMS = f"{BASE}/202601/clms_202601.csv"
MBR  = f"{BASE}/202601/mbr_202601.csv"

CLOSE_DATE = dt.datetime(2026, 1, 31)

# ---- Load CSVs
clms = pd.read_csv(CLMS, parse_dates=['incurred_year_month'])
mbr  = pd.read_csv(MBR, parse_dates=['year_month'])
print(f"clms rows: {len(clms)} | mbr rows: {len(mbr)}")

# ---- Open workbook
wb = load_workbook(JAN_FILE)
print("Loaded JAN file, sheets:", wb.sheetnames)

# ---- Update controls
wb['controls']['B5'] = CLOSE_DATE

# ---- Snapshot a template row (claims row 7 formulas E:I, mbr row 7 formulas D:E)
claims_ws = wb['claims']
mbr_ws = wb['mbr']

# Capture template cell styles and formulas from row 7 of claims
claims_templates = {col: claims_ws.cell(row=7, column=col) for col in range(1, 10)}
mbr_templates = {col: mbr_ws.cell(row=7, column=col) for col in range(1, 6)}

# ---- Clear existing data rows on claims/mbr
def clear_rows_from(ws, start_row):
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None

clear_rows_from(claims_ws, 7)
clear_rows_from(mbr_ws, 7)

# ---- Write new claims data with formulas in cols E:I
# claims columns: A incurred_year_month, B plan_type, C claim_type, D plan_paid,
# E year formula, F month formula, G lag, H CF (VLOOKUP), I ultimate = D/H
for i, row in enumerate(clms.itertuples(index=False), start=7):
    claims_ws.cell(row=i, column=1, value=row.incurred_year_month.to_pydatetime())
    claims_ws.cell(row=i, column=2, value=row.plan_type)
    claims_ws.cell(row=i, column=3, value=row.claim_type)
    claims_ws.cell(row=i, column=4, value=float(row.plan_paid))
    claims_ws.cell(row=i, column=5, value=f"=YEAR(A{i})")
    claims_ws.cell(row=i, column=6, value=f"=MONTH(A{i})")
    claims_ws.cell(row=i, column=7, value=f"=(YEAR(close_date)-YEAR(A{i}))*12+MONTH(close_date)-MONTH(A{i})")
    claims_ws.cell(row=i, column=8, value=f"=VLOOKUP(B{i}&C{i}&G{i},cf!D:E,2,FALSE)")
    claims_ws.cell(row=i, column=9, value=f"=D{i}/H{i}")
    # Copy styling from template row 7
    for col in range(1, 10):
        src = claims_templates[col]
        dst = claims_ws.cell(row=i, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

# Set A column as date format if needed
for i in range(7, 7 + len(clms)):
    claims_ws.cell(row=i, column=1).number_format = 'm/d/yyyy'

# ---- Write new mbr data
# mbr columns: A year_month, B plan_type, C member_months, D year, E month
for i, row in enumerate(mbr.itertuples(index=False), start=7):
    mbr_ws.cell(row=i, column=1, value=row.year_month.to_pydatetime())
    mbr_ws.cell(row=i, column=2, value=row.plan_type)
    mbr_ws.cell(row=i, column=3, value=int(row.member_months))
    mbr_ws.cell(row=i, column=4, value=f"=YEAR(A{i})")
    mbr_ws.cell(row=i, column=5, value=f"=MONTH(A{i})")
    for col in range(1, 6):
        src = mbr_templates[col]
        dst = mbr_ws.cell(row=i, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format

for i in range(7, 7 + len(mbr)):
    mbr_ws.cell(row=i, column=1).number_format = 'm/d/yyyy'

# ---- Read completion factors from cf tab
cf_ws = wb['cf']
cf_map = {}
for r in cf_ws.iter_rows(min_row=7, values_only=True):
    if r[0] is None: continue
    plan, claim, lag, key, factor = r[0], r[1], r[2], r[3], r[4]
    if factor is not None:
        cf_map[(plan, claim, int(lag))] = float(factor)
print(f"CF entries: {len(cf_map)}")

# ---- Compute lag, CF, ultimate in Python (mirror the sheet formulas)
def compute_lag(close_dt, inc_dt):
    return (close_dt.year - inc_dt.year) * 12 + (close_dt.month - inc_dt.month)

clms['lag'] = clms['incurred_year_month'].apply(lambda d: compute_lag(CLOSE_DATE, d))
clms['cf'] = clms.apply(lambda r: cf_map[(r['plan_type'], r['claim_type'], int(min(r['lag'], 23)))], axis=1)
clms['ultimate'] = clms['plan_paid'] / clms['cf']
clms['year'] = clms['incurred_year_month'].dt.year
clms['month'] = clms['incurred_year_month'].dt.month

mbr['year'] = mbr['year_month'].dt.year
mbr['month'] = mbr['year_month'].dt.month

# ---- Update pivot cached values for 2024 and 2025
pivot_ws = wb['pivot']

# mbr pivot: row 4 = 2024, row 5 = 2025, row 6 = Grand Total. Cols D..O = months 1..12, P = Grand Total
mbr_yr_mth = mbr.groupby(['year','month'])['member_months'].sum().unstack(fill_value=0)
for m in range(1, 13):
    for yr, row_idx in [(2024, 4), (2025, 5)]:
        val = int(mbr_yr_mth.loc[yr, m]) if (yr in mbr_yr_mth.index and m in mbr_yr_mth.columns) else 0
        pivot_ws.cell(row=row_idx, column=3 + m, value=val if val > 0 else None)
# Grand totals
for yr, row_idx in [(2024, 4), (2025, 5)]:
    if yr in mbr_yr_mth.index:
        pivot_ws.cell(row=row_idx, column=16, value=int(mbr_yr_mth.loc[yr].sum()))
# Grand Total row (row 6)
for m in range(1, 13):
    s = 0
    for yr in [2024, 2025]:
        if yr in mbr_yr_mth.index and m in mbr_yr_mth.columns:
            s += int(mbr_yr_mth.loc[yr, m])
    pivot_ws.cell(row=6, column=3 + m, value=s if s > 0 else None)
pivot_ws.cell(row=6, column=16, value=int(mbr_yr_mth.loc[[y for y in [2024,2025] if y in mbr_yr_mth.index]].values.sum()))

# clms pivot: row 22 = 2024, row 23 = 2025, row 24 = Grand Total
clms_yr_mth = clms.groupby(['year','month'])['ultimate'].sum().unstack(fill_value=0)
for m in range(1, 13):
    for yr, row_idx in [(2024, 22), (2025, 23)]:
        val = float(clms_yr_mth.loc[yr, m]) if (yr in clms_yr_mth.index and m in clms_yr_mth.columns) else 0
        pivot_ws.cell(row=row_idx, column=3 + m, value=val if val > 0 else None)
for yr, row_idx in [(2024, 22), (2025, 23)]:
    if yr in clms_yr_mth.index:
        pivot_ws.cell(row=row_idx, column=16, value=float(clms_yr_mth.loc[yr].sum()))
for m in range(1, 13):
    s = 0.0
    for yr in [2024, 2025]:
        if yr in clms_yr_mth.index and m in clms_yr_mth.columns:
            s += float(clms_yr_mth.loc[yr, m])
    pivot_ws.cell(row=24, column=3 + m, value=s if s > 0 else None)
pivot_ws.cell(row=24, column=16,
              value=float(clms_yr_mth.loc[[y for y in [2024,2025] if y in clms_yr_mth.index]].values.sum()))

# ---- Rebuild combined data (mbr records first, then claims records)
cd_ws = wb['combined data']
# Use delete_rows to remove stale data cleanly
if cd_ws.max_row > 6:
    cd_ws.delete_rows(idx=7, amount=cd_ws.max_row - 6)
print(f"  combined data cleared, max_row now: {cd_ws.max_row}")

row = 7
# mbr records: year, month, plan_type, (blank), member_months, (blank)
mbr_sorted = mbr.sort_values(['year','month','plan_type']).reset_index(drop=True)
print(f"  mbr rows to write: {len(mbr_sorted)}")
for i, rec in enumerate(mbr_sorted.itertuples(index=False)):
    target = 7 + i
    cd_ws.cell(row=target, column=1, value=int(rec.year))
    cd_ws.cell(row=target, column=2, value=int(rec.month))
    cd_ws.cell(row=target, column=3, value=rec.plan_type)
    cd_ws.cell(row=target, column=5, value=int(rec.member_months))

mbr_end = 7 + len(mbr_sorted) - 1
clms_start = mbr_end + 1
print(f"  mbr rows 7..{mbr_end}, clms starts at {clms_start}")

# claims records: year, month, plan_type, claim_type, (blank), ultimate_incurred
order = {'HMO':0, 'PPO':1, 'IP':0, 'Non-IP':1}
clms_sorted = clms.sort_values(
    by=['year','month','plan_type','claim_type'],
    key=lambda s: s.map(order) if s.name in ('plan_type','claim_type') else s
).reset_index(drop=True)
print(f"  clms rows to write: {len(clms_sorted)}")
for i, rec in enumerate(clms_sorted.itertuples(index=False)):
    target = clms_start + i
    cd_ws.cell(row=target, column=1, value=int(rec.year))
    cd_ws.cell(row=target, column=2, value=int(rec.month))
    cd_ws.cell(row=target, column=3, value=rec.plan_type)
    cd_ws.cell(row=target, column=4, value=rec.claim_type)
    cd_ws.cell(row=target, column=6, value=float(rec.ultimate))

total = len(mbr_sorted) + len(clms_sorted)
print(f"  combined data total records written: {total}")

# ---- Save
wb.save(JAN_FILE)
print(f"Saved: {JAN_FILE}")
